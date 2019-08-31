#! blankRowInserter.py
# You can write this program by reading in the contents of the spreadsheet.
# Then, when writing out the new spreadsheet, use a for loop to copy the first N lines. 
# For the remaining lines, add M to the row number in the output spreadsheet.

import openpyxl
import sys
import os



N  = int(sys.argv[1]) # Insert blank row after this row
M = int(sys.argv[2]) #Number of rows to insert

Filename = input('Please enter the new file name (xlsx).')

wb = openpyxl.Workbook()
dest_filename = Filename + '.xlsx'

ws1 = wb.active
ws1.title = "Updated"

print('Opening workbook...')
sourcefilename = os.path.basename(sys.argv[3]) # cwd folder/filename

example = openpyxl.load_workbook(sourcefilename)

sheet = example.active

print('Writing to workbook: %s' %(dest_filename))

for i in range(1, sheet.max_row):
    for j in range(1, sheet.max_column+1):
        ws1.cell(row=i,column=j).value = sheet.cell(row=i,column=j).value

ws1.insert_rows(N+1,M) #insert after the row specified.
print('Number of rows written: %s' %(ws1.max_row))
wb.save(filename = dest_filename)

print('Done')
