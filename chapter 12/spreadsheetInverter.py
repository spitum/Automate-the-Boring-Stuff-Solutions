import openpyxl
import sys
import os

#Excel has limitations on number of columns. Please try this with a temp[:1000] or less first if the source file is very big.

# TODO : convert to a function with source file name as argument. 

print('Opening workbook...')

sourcefilename = os.path.basename(sys.argv[1])

sourcefilename_split = sourcefilename.split('.')

wb = openpyxl.load_workbook(sourcefilename)

ws1 = wb.active


new = openpyxl.Workbook()
dest_filename = f'Pivoted_{sourcefilename_split[0]}.xlsx'

ws2 = new.active
ws2.title = "Updated"


temp = []
for row in ws1.rows:
    # Append a list of column values
    temp.append( [cell.value for cell in row] )
print(len(temp))

for i in range(len(temp[:100])):
    ws2.cell(row=1,column=i+1).value = temp[i][0] #first row
    for j in range(len(temp[i])):
        ws2.cell(row=j+1,column=i+1).value = temp[i][j] #data


new.save(filename = dest_filename)

print('Done')
