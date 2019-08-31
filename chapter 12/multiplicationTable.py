from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font
import sys
import os

wb = Workbook()


ws1 = wb.active
ws1.title = "range names"

fontObj1  = Font(size = 12, bold = True)

N = int(sys.argv[1])

for row in range(1,N+1):
    ws1.cell(row=row+1,column=1).value = row
    ws1.cell(row=1,column=row+1).value = row
    ws1.cell(row = row+1 , column = 1).font = fontObj1
    ws1.cell(row = 1 , column = row+1).font = fontObj1


for colNum in range(1,ws1.max_column):
    for rowNum in range(1,ws1.max_row):
        ws1.cell(row = rowNum+1, column = colNum+1).value = colNum * rowNum

wb.save(filename = 'multiplationTable_%s.xlsx' %(N))
